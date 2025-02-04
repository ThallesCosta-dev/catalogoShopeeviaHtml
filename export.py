from bs4 import BeautifulSoup
import pandas as pd
import logging
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from PIL import Image as PILImage
from datetime import datetime

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def download_and_resize_image(url):
    try:
        response = requests.get(url)
        img = PILImage.open(BytesIO(response.content))
        
        # Redimensionar mantendo proporção
        basewidth = 100
        wpercent = (basewidth/float(img.size[0]))
        hsize = int((float(img.size[1])*float(wpercent)))
        img = img.resize((basewidth, hsize), PILImage.Resampling.LANCZOS)
        
        # Converter para PNG
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        return img_byte_arr
    except Exception as e:
        logger.error(f"Erro ao processar imagem: {str(e)}")
        return None

def extrair_dados_produto(item):
    try:
        # Imagem e nome
        img_element = item.find('img', class_='object-contain')
        nome = img_element.get('alt', 'Nome não disponível')
        imagem = img_element.get('src', '')
        
        # Valor
        valor_element = item.find('span', class_='text-base/5')
        valor = valor_element.text.strip() if valor_element else "Valor não disponível"
        # Converter valor para número
        if valor != "Valor não disponível":
            valor = float(valor.replace('R$', '').replace('.', '').replace(',', '.').strip())
        
        # Vendas
        vendas_element = item.find('div', class_='truncate text-shopee-black87 text-xs min-h-4')
        vendas = vendas_element.text.strip() if vendas_element else "0 vendidos"
        
        # Converter vendas para número, tratando valores com 'mil'
        if 'mil' in vendas.lower():
            # Remove todos os caracteres não numéricos exceto vírgula e ponto
            numero = ''.join(c for c in vendas if c.isdigit() or c in '.,')
            # Substitui vírgula por ponto para conversão
            numero = numero.replace(',', '.')
            # Multiplica por 1000 e converte para inteiro
            vendas = int(float(numero) * 1000)
        else:
            # Para números sem 'mil', apenas extrai os dígitos
            vendas = int(''.join(filter(str.isdigit, vendas))) if vendas != "0 vendidos" else 0
        
        # Link
        link = item.get('href', '')
        if link and not link.startswith('http'):
            link = f"https://shopee.com.br{link}"
        
        # Avaliação
        avaliacao_element = item.find('div', class_='text-shopee-black87 text-xs/sp14 flex-none')
        avaliacao = avaliacao_element.text.strip() if avaliacao_element else "0"
        # Converter avaliação para número
        avaliacao = float(avaliacao) if avaliacao != "Sem avaliação" else 0
        
        return {
            'Nome do Produto': nome,
            'Valor (R$)': valor,
            'Quantidade Vendida': vendas,
            'Avaliação (⭐)': avaliacao,
            'Link do Produto': link,
            'URL da Imagem': imagem
        }
    except Exception as e:
        logger.error(f"Erro ao extrair dados do produto: {str(e)}")
        return None

def formatar_planilha(workbook_path, produtos):
    try:
        wb = Workbook()
        sheet = wb.active
        
        # Configurar altura da linha para acomodar as imagens
        sheet.row_dimensions[1].height = 40  # Altura do cabeçalho
        
        # Adicionar cabeçalhos
        headers = ['Imagem', 'Nome do Produto', 'Valor (R$)', 'Quantidade Vendida', 'Avaliação (⭐)', 'Link do Produto']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adicionar dados e imagens
        for row_idx, produto in enumerate(produtos, 2):
            # Configurar altura da linha para imagens
            sheet.row_dimensions[row_idx].height = 75
            
            # Tentar baixar e inserir imagem
            img_data = download_and_resize_image(produto['URL da Imagem'])
            if img_data:
                img = Image(img_data)
                img.width = 100  # Largura fixa em pixels
                img.height = 75  # Altura fixa em pixels
                sheet.add_image(img, f'A{row_idx}')
            
            # Adicionar outros dados
            sheet.cell(row=row_idx, column=2, value=produto['Nome do Produto'])
            sheet.cell(row=row_idx, column=3, value=produto['Valor (R$)'])
            sheet.cell(row=row_idx, column=4, value=produto['Quantidade Vendida'])
            sheet.cell(row=row_idx, column=5, value=produto['Avaliação (⭐)'])
            sheet.cell(row=row_idx, column=6, value=produto['Link do Produto'])
            
            # Formatar células
            for col in range(2, 7):
                cell = sheet.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Ajustar largura das colunas
        sheet.column_dimensions['A'].width = 15  # Coluna de imagens
        for col in range(2, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 100)
            sheet.column_dimensions[column].width = adjusted_width
        
        # Salvar planilha
        wb.save(workbook_path)
        logger.info(f"Planilha formatada salva em: {workbook_path}")
        
    except Exception as e:
        logger.error(f"Erro ao formatar planilha: {str(e)}")

def processar_arquivo_html(arquivo):
    try:
        logger.info(f"Processando arquivo: {arquivo}")
        with open(arquivo, 'r', encoding='utf-8') as f:
            conteudo = f.read()
        
        soup = BeautifulSoup(conteudo, 'html.parser')
        items = soup.find_all('a', class_='contents')
        
        produtos = []
        for item in items:
            produto = extrair_dados_produto(item)
            if produto:
                produtos.append(produto)
                logger.info(f"Produto extraído: {produto['Nome do Produto'][:30]}...")
        
        return produtos
    except Exception as e:
        logger.error(f"Erro ao processar arquivo {arquivo}: {str(e)}")
        return []

if __name__ == "__main__":
    try:
        # Encontrar todos os arquivos HTML que começam com 'zem'
        arquivos_html = glob.glob('zem*.html')
        
        if not arquivos_html:
            logger.error("Nenhum arquivo HTML encontrado com o padrão 'zem*.html'")
            exit()
        
        todos_produtos = []
        
        # Processar cada arquivo
        for arquivo in arquivos_html:
            produtos = processar_arquivo_html(arquivo)
            todos_produtos.extend(produtos)
        
        if todos_produtos:
            # Converter para DataFrame
            df = pd.DataFrame(todos_produtos)
            
            # Ordenar por quantidade de vendas (decrescente) antes de remover duplicatas
            df = df.sort_values('Quantidade Vendida', ascending=False)
            
            # Remover duplicatas mantendo a primeira ocorrência (que será a com mais vendas)
            df_sem_duplicados = df.drop_duplicates(subset=['Nome do Produto'], keep='first')
            
            # Converter de volta para lista de dicionários
            todos_produtos = df_sem_duplicados.to_dict('records')
            
            # Criar e formatar planilha
            output_file = 'produtos_shopee.xlsx'
            formatar_planilha(output_file, todos_produtos)
            
            logger.info(f"Total de {len(todos_produtos)} produtos únicos salvos no XLSX")
            logger.info("Produtos duplicados foram removidos, mantendo os anúncios com mais vendas")
        else:
            logger.error("Nenhum produto encontrado nos arquivos")
            
    except Exception as e:
        logger.error(f"Erro na execução principal: {str(e)}")